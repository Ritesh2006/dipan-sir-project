"""Google Drive & Sheets Integration Manager with Local Excel/Disk Fallback."""

import os
import shutil
import datetime
from pathlib import Path
from typing import Dict, Any, Tuple, Optional
import openpyxl

from app.config.settings import settings
from app.excel.formatter import format_header_row, auto_fit_columns
from app.utils.logger import logger

# Output Master Spreadsheet Name
EXHIBITION_EXCEL_PATH = settings.BASE_DIR / "data" / "exports" / "National Exhibition 2026 – Data.xlsx"

TAB_HEADERS = {
    "Stall Data": [
        "Submission ID", "Timestamp", "Stall Name", "Stall No.", "Organization",
        "Category", "Person", "Designation", "Audio Drive Link", "Image Drive Link",
        "Brochure Drive Link", "Transcript", "Verification Status"
    ],
    "Science Exhibition Data": [
        "Submission ID", "Timestamp", "Exhibit/Project Name", "Stall No.", "Organization/Institution",
        "Category", "Presenter", "Designation/Class", "Audio Drive Link", "Image Drive Link",
        "Brochure Drive Link", "Transcript", "Verification Status"
    ],
    "Live Lecture Data": [
        "Submission ID", "Timestamp", "Lecture Title", "Speaker", "Designation",
        "Organization", "Topic/Category", "Date/Time", "Audio Drive Link", "Image Drive Link",
        "Brochure Drive Link", "Transcript", "Verification Status"
    ]
}


class GoogleExhibitionService:
    """Manages Google Drive & Sheets integration with automatic local fallback."""

    def __init__(self, excel_path: Optional[Path] = None):
        self.excel_path = excel_path or EXHIBITION_EXCEL_PATH
        self.excel_path.parent.mkdir(parents=True, exist_ok=True)
        self.ensure_sheets_exist()

    def ensure_sheets_exist(self) -> None:
        """Create master workbook with standard sheets (Stall Data, Science Exhibition Data, Live Lecture Data)."""
        if not self.excel_path.exists():
            wb = openpyxl.Workbook()
            # Remove default sheet
            default_sheet = wb.active
            
            for tab_name, headers in TAB_HEADERS.items():
                ws = wb.create_sheet(title=tab_name)
                ws.append(headers)
                format_header_row(ws)
                auto_fit_columns(ws)

            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            wb.save(str(self.excel_path))
            wb.close()
            logger.info(f"Initialized National Exhibition Master Excel at {self.excel_path}")

    def append_submission_row(self, tab_name: str, row_dict: Dict[str, Any]) -> bool:
        """Append ONE new row to the specified sheet."""
        self.ensure_sheets_exist()
        try:
            wb = openpyxl.load_workbook(str(self.excel_path))
            if tab_name not in wb.sheetnames:
                ws = wb.create_sheet(title=tab_name)
                headers = TAB_HEADERS.get(tab_name, list(row_dict.keys()))
                ws.append(headers)
                format_header_row(ws)
            else:
                ws = wb[tab_name]

            headers = [str(cell.value or "") for cell in ws[1]]
            row_vals = [row_dict.get(h, "") for h in headers]

            ws.append(row_vals)
            auto_fit_columns(ws)
            wb.save(str(self.excel_path))
            wb.close()
            logger.info(f"Appended submission '{row_dict.get('Submission ID')}' row to sheet '{tab_name}'")

            # Automatic Google Drive Sync of Master Excel file
            try:
                from app.integrations.google_drive_uploader import GoogleDriveUploader
                GoogleDriveUploader.upload_excel_file(self.excel_path)
            except Exception as sync_err:
                logger.warning(f"Excel Google Drive sync notice: {sync_err}")

            return True
        except Exception as e:
            logger.error(f"Failed to append row to {tab_name}: {e}")
            return False

    def update_submission_row(self, tab_name: str, submission_id: str, updates: Dict[str, Any]) -> bool:
        """Update existing row in Google Sheets / Excel by Submission ID."""
        self.ensure_sheets_exist()
        try:
            wb = openpyxl.load_workbook(str(self.excel_path))
            if tab_name not in wb.sheetnames:
                wb.close()
                return False

            ws = wb[tab_name]
            headers = [str(cell.value or "") for cell in ws[1]]
            id_col_idx = headers.index("Submission ID") + 1 if "Submission ID" in headers else 1

            target_row = None
            for r in range(2, ws.max_row + 1):
                val = str(ws.cell(row=r, column=id_col_idx).value or "")
                if val.strip() == submission_id.strip():
                    target_row = r
                    break

            if target_row is None:
                logger.warning(f"Submission ID '{submission_id}' not found in sheet '{tab_name}' for update.")
                wb.close()
                return False

            for key, new_val in updates.items():
                if key in headers:
                    col_idx = headers.index(key) + 1
                    ws.cell(row=target_row, column=col_idx, value=new_val)

            auto_fit_columns(ws)
            wb.save(str(self.excel_path))
            wb.close()
            logger.info(f"Updated Submission ID '{submission_id}' in sheet '{tab_name}' with fields: {list(updates.keys())}")
            return True
        except Exception as e:
            logger.error(f"Failed to update row for Submission ID '{submission_id}': {e}")
            return False
