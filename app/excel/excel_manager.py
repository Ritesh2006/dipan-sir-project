"""Excel Manager for creating, appending, and updating voice log spreadsheets."""

import os
import shutil
import threading
import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.config.settings import settings
from app.excel.formatter import format_header_row, auto_fit_columns
from app.utils.validators import sanitize_excel_value
from app.utils.logger import logger


DEFAULT_HEADERS = ["Speaker / Dignitary", "Topic / Agenda", "Decision / Action", "Amount / Budget", "Date", "Logged At"]


class ExcelManager:
    """Manages offline Excel logging operations safely using openpyxl."""

    def __init__(self, excel_path: Optional[Path] = None, sheet_name: str = "VIP Meeting Logs"):
        self.excel_path = excel_path or settings.EXCEL_FILE
        self.sheet_name = sheet_name
        self._lock = threading.RLock()
        self.excel_path.parent.mkdir(parents=True, exist_ok=True)
        self.ensure_file_exists()

    def ensure_file_exists(self) -> None:
        """Create workbook with headers if file does not exist."""
        with self._lock:
            if not self.excel_path.exists():
                try:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = self.sheet_name
                    ws.append(DEFAULT_HEADERS)
                    format_header_row(ws)
                    auto_fit_columns(ws)
                    wb.save(str(self.excel_path))
                    wb.close()
                    logger.info(f"Created new Excel file at '{self.excel_path}'.")
                except Exception as e:
                    logger.error(f"Failed to create Excel file at '{self.excel_path}': {e}")

    def log_record(self, record: Any) -> Path:
        """Log VipMeetingRecord or dict to Excel workbook."""
        data_dict = record.to_excel_row() if hasattr(record, "to_excel_row") else (
            record.model_dump() if hasattr(record, "model_dump") else (
                record.dict() if hasattr(record, "dict") else record
            )
        )
        with self._lock:
            self.ensure_file_exists()
            try:
                wb = openpyxl.load_workbook(str(self.excel_path))
                ws = wb[self.sheet_name] if self.sheet_name in wb.sheetnames else wb.active

                headers = [str(cell.value or "") for cell in ws[1]] if ws.max_row >= 1 else DEFAULT_HEADERS

                row_vals = []
                for h in headers:
                    if h == "Logged At":
                        row_vals.append(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    else:
                        row_vals.append(sanitize_excel_value(data_dict.get(h, "")))

                ws.append(row_vals)
                auto_fit_columns(ws)
                wb.save(str(self.excel_path))
                wb.close()
                logger.info(f"Logged VIP record to Excel: {data_dict}")
            except Exception as e:
                logger.error(f"Failed to log record to Excel: {e}")
        return self.excel_path

    def append_record(self, record: Any) -> Tuple[bool, str]:
        """Append record to Excel workbook and return (success, message) tuple."""
        try:
            path = self.log_record(record)
            return True, f"Successfully logged record to Excel file '{path.name}'."
        except Exception as e:
            err_msg = f"Failed to log record to Excel: {e}"
            logger.error(err_msg)
            return False, err_msg

