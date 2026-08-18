"""Excel data validation and duplicate detection module."""

from typing import List, Dict, Any
from openpyxl.worksheet.worksheet import Worksheet
from app.utils.logger import logger


def is_duplicate_record(ws: Worksheet, new_data: Dict[str, Any], headers: List[str]) -> bool:
    """Check if new_data is already present in recent rows of the worksheet.

    Matches based on unique combinations (e.g., Name + Roll + Date).
    """
    if ws.max_row <= 1:
        return False

    name = str(new_data.get("Name", "")).strip().lower()
    roll = str(new_data.get("Roll", "")).strip()
    date = str(new_data.get("Date", "")).strip()

    if not name and not roll:
        return False

    header_indices = {h: i + 1 for i, h in enumerate(headers)}
    name_idx = header_indices.get("Name")
    roll_idx = header_indices.get("Roll")
    date_idx = header_indices.get("Date")

    # Scan last 100 rows for matching record
    min_row = max(2, ws.max_row - 100)
    for row in range(ws.max_row, min_row - 1, -1):
        r_name = str(ws.cell(row=row, column=name_idx).value or "").strip().lower() if name_idx else ""
        r_roll = str(ws.cell(row=row, column=roll_idx).value or "").strip() if roll_idx else ""
        r_date = str(ws.cell(row=row, column=date_idx).value or "").strip() if date_idx else ""

        if name and roll and date:
            if r_name == name and r_roll == roll and r_date == date:
                logger.info(f"Duplicate entry detected: Name='{name}', Roll='{roll}', Date='{date}' in row {row}.")
                return True
        elif name and roll:
            if r_name == name and r_roll == roll:
                logger.info(f"Duplicate entry detected: Name='{name}', Roll='{roll}' in row {row}.")
                return True

    return False
