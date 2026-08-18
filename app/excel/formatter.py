"""Excel formatting helper using openpyxl."""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet


def format_header_row(ws: Worksheet) -> None:
    """Format row 1 with modern header styling and auto-filters."""
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="medium", color="1F4E79"),
    )

    ws.row_dimensions[1].height = 26

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def auto_fit_columns(ws: Worksheet) -> None:
    """Adjust column widths dynamically based on max text length."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or "")
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
